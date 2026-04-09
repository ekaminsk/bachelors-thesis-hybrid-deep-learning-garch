#!/usr/bin/env python3
"""
Thesis Data Aggregator  -  5-minute level
==========================================
Merges ALL data sources into a single time-indexed file at 5-min resolution.
Time index: window_end (UTC) -- the candle/snapshot closes at this timestamp.
  e.g. window_end = 2026-03-06 00:05:00  covers  00:00:00 -> 00:05:00

Sources:
  DEX - Uniswap v3 USDC/USDT 0.01%, Ethereum mainnet
    klines_dex/DD-MM-klines.csv          5-min OHLCV  (window_start -> +5min = window_end)
    klines_dex/DD-MM-swaps-raw.csv       individual swaps  (unix seconds -> floored to 5-min)
    liquidity_state/DD-MM-pool.csv       5-min pool state  (window_end)
    liquidity_flow/DD-MM-ticks.csv       5-min tick dist.  (window_end)
    mints_burns/DD-MM-mints-burns.csv    5-min LP events   (window_end)
  CEX - Binance USDC/USDT spot
    klines_cex/USDCUSDT-5m-YYYY-MM-DD.csv   5-min OHLCV  (unix microseconds -> window_end)
    orderbook/DD-MM-orderbook.csv            5-min OB snap (window_end)
  DUNE - Ethereum on-chain (Dune Analytics exports)  [SKELETON -- fill paths when ready]
    dune_ONCHAIN/q3_gas_price.csv          Query 6763557: Gas Price Time Series
    dune_ONCHAIN/q4_gas_used.csv           Query 6763559: Gas Used per Block & Utilization
    dune_ONCHAIN/q5_mempool_congestion.csv Query 6763560: Mempool Congestion Proxies
    dune_ONCHAIN/q6_supply_changes.csv     Query 6763561: USDC/USDT True Mints & Burns
    dune_ONCHAIN/q2_cex_flows.csv          Query 6763555: CEX Inflows & Outflows
    dune_ONCHAIN/q1_whale_transfers.csv    Query 6763552: Whale Transfers

Column prefix convention:
  dex_klines_*   DEX 5-min OHLCV (price, volume, imbalance)
  dex_swaps_*    DEX raw swap aggregates (wallet counts, tick range)
  dex_pool_*     DEX pool state (TVL, liquidity, sqrtPrice-derived price)
  dex_ticks_*    DEX tick liquidity distribution
  dex_lp_*       DEX LP mint/burn events
  cex_price_*    CEX Binance price OHLC
  cex_volume_*   CEX Binance volume
  cex_ob_*       CEX Binance orderbook
  dune_gas_*     On-chain gas metrics (Dune)
  dune_mempool_* On-chain mempool congestion proxies (Dune)
  dune_supply_*  USDC/USDT on-chain mint/burn supply changes (Dune)
  dune_flows_*   On-chain CEX inflow/outflow (Dune)
  dune_whale_*   Whale transfer aggregates (Dune)

Usage:
    python aggregate_5min.py

Output:
    D:/data/thesis_5min.xlsx
    D:/data/thesis_5min.csv
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================================================================
# CONFIG
# ==============================================================================

DEX_DIR   = Path("D:/data/uniswap_DEX")
CEX_DIR   = Path("D:/data/binance_CEX")
DUNE_DIR  = Path("D:/data/dune_ONCHAIN")   # drop Dune CSV exports here
OUT_XLSX  = Path("D:/data/thesis_5min.xlsx")
OUT_CSV   = Path("D:/data/thesis_5min.csv")

FIVE_MIN  = timedelta(minutes=5)
YEAR      = "2026"

# Set True once you have placed Dune CSV exports in DUNE_DIR
DUNE_ENABLED = True

# ==============================================================================
# FILENAME HELPERS
# ==============================================================================

def dex_stem(date_str):
    """'2026-03-06' -> '06-03'  (DEX filename prefix)"""
    return f"{date_str[8:10]}-{date_str[5:7]}"

def _dates_from_glob(folder, pattern, parse_fn):
    dates = set()
    for f in folder.glob(pattern):
        result = parse_fn(f)
        if result:
            dates.add(result)
    return sorted(dates)

def dex_kline_dates():
    return _dates_from_glob(
        DEX_DIR / "klines_dex", "*-klines.csv",
        lambda f: f"{YEAR}-{f.stem.split('-')[1]}-{f.stem.split('-')[0]}"
    )

def cex_kline_dates():
    return _dates_from_glob(
        CEX_DIR / "klines_cex", "USDCUSDT-5m-*.csv",
        lambda f: "-".join(f.stem.split("-")[2:5])
    )

def orderbook_dates():
    return _dates_from_glob(
        CEX_DIR / "orderbook", "*-orderbook.csv",
        lambda f: f"{YEAR}-{f.stem.split('-')[1]}-{f.stem.split('-')[0]}"
    )

def pool_dates():
    return _dates_from_glob(
        DEX_DIR / "liquidity_state", "*-pool.csv",
        lambda f: f"{YEAR}-{f.stem.split('-')[1]}-{f.stem.split('-')[0]}"
    )

def ticks_dates():
    return _dates_from_glob(
        DEX_DIR / "liquidity_flow", "*-ticks.csv",
        lambda f: f"{YEAR}-{f.stem.split('-')[1]}-{f.stem.split('-')[0]}"
    )

def mb_dates():
    return _dates_from_glob(
        DEX_DIR / "mints_burns", "*-mints-burns.csv",
        lambda f: f"{YEAR}-{f.stem.split('-')[1]}-{f.stem.split('-')[0]}"
    )

# ==============================================================================
# DEX LOADERS
# ==============================================================================

def load_all_dex_klines():
    """
    5-min OHLCV aggregated from raw swaps.
    Source uses window_start; we add 5min to get window_end.
    imbalance = net_amount0 / |amount0|:
      +1 = all flow was USDC into pool (swap: USDC -> USDT, i.e. buying USDT)
      -1 = all flow was USDC out of pool (swap: USDT -> USDC, i.e. buying USDC)
    """
    frames = []
    for date_str in dex_kline_dates():
        path = DEX_DIR / "klines_dex" / f"{dex_stem(date_str)}-klines.csv"
        df = pd.read_csv(path, parse_dates=["window_start"])
        df["window_end"] = df["window_start"] + FIVE_MIN
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end")
    out = out.set_index("window_end").drop(columns=["window_start"])
    out.columns = [
        "dex_klines_price_open",
        "dex_klines_price_high",
        "dex_klines_price_low",
        "dex_klines_price_close",
        "dex_klines_volume_usd",
        "dex_klines_n_swaps",
        "dex_klines_imbalance",
        "dex_klines_large_trades_count",
        "dex_klines_large_trades_usd",
    ]
    return out


def load_all_dex_swaps_raw():
    """
    Individual swap events aggregated to 5-min buckets.
    Adds wallet-level and tick-range info not captured in klines.
    """
    frames = []
    for date_str in dex_kline_dates():
        path = DEX_DIR / "klines_dex" / f"{dex_stem(date_str)}-swaps-raw.csv"
        if not path.exists():
            continue
        df = pd.read_csv(path)
        df["amountUSD"] = pd.to_numeric(df["amountUSD"], errors="coerce").abs()
        df["dt"] = pd.to_datetime(
            df["timestamp"].astype(int), unit="s", utc=True
        ).dt.tz_localize(None)
        df["window_end"] = df["dt"].dt.floor("5min") + FIVE_MIN
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    all_swaps = pd.concat(frames, ignore_index=True)
    agg = all_swaps.groupby("window_end").agg(
        dex_swaps_n_total            =("amountUSD", "count"),
        dex_swaps_unique_senders     =("sender",    "nunique"),
        dex_swaps_unique_recipients  =("recipient", "nunique"),
        dex_swaps_tick_min           =("tick",      "min"),
        dex_swaps_tick_max           =("tick",      "max"),
        dex_swaps_usd_mean_per_swap  =("amountUSD", "mean"),
        dex_swaps_usd_median_per_swap=("amountUSD", "median"),
        dex_swaps_usd_max_single_swap=("amountUSD", "max"),
    )
    return agg


def load_all_dex_pool():
    """
    5-min pool state: sqrtPrice-derived price, TVL, active liquidity.
    dex_pool_price  = USDC/USDT from sqrtPriceX96 (should be ~1.0 for stablecoin pair)
    dex_pool_tvl_usd = total value locked in both tokens combined
    """
    frames = []
    for date_str in pool_dates():
        path = DEX_DIR / "liquidity_state" / f"{dex_stem(date_str)}-pool.csv"
        df = pd.read_csv(path, parse_dates=["window_end"])
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end").set_index("window_end")
    for col in ["price", "token0_price", "token1_price", "tvl_usd", "tvl_token0", "tvl_token1"]:
        out[col] = pd.to_numeric(out[col], errors="coerce")
    out = out.rename(columns={
        "liquidity":    "dex_pool_liquidity",
        "tick":         "dex_pool_tick",
        "sqrt_price":   "dex_pool_sqrt_price",
        "price":        "dex_pool_price",
        "token0_price": "dex_pool_token0_price",
        "token1_price": "dex_pool_token1_price",
        "tvl_token0":   "dex_pool_tvl_usdc",
        "tvl_token1":   "dex_pool_tvl_usdt",
        "tvl_usd":      "dex_pool_tvl_usd",
    })
    return out[[c for c in out.columns if c.startswith("dex_pool_")]]


def load_all_dex_ticks():
    """
    5-min tick liquidity distribution.
    net_liq_above/below shows how much liquidity is positioned above vs below
    the current price -- useful for understanding LP range concentration.
    """
    frames = []
    for date_str in ticks_dates():
        path = DEX_DIR / "liquidity_flow" / f"{dex_stem(date_str)}-ticks.csv"
        df = pd.read_csv(path, parse_dates=["window_end"])
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end").set_index("window_end")
    out = out.rename(columns={
        "current_tick":    "dex_ticks_current_tick",
        "tick_lo":         "dex_ticks_active_lo",
        "tick_hi":         "dex_ticks_active_hi",
        "n_ticks":         "dex_ticks_n_active",
        "total_liq_gross": "dex_ticks_total_liq_gross",
        "net_liq_above":   "dex_ticks_net_liq_above",
        "net_liq_below":   "dex_ticks_net_liq_below",
    })
    return out


def load_all_dex_mints_burns():
    """
    5-min LP position events.
    Mints = LP adds liquidity (capital inflow to pool).
    Burns = LP removes liquidity (capital outflow from pool).
    These are NOT swap events -- they change pool depth, not price.
    """
    frames = []
    for date_str in mb_dates():
        path = DEX_DIR / "mints_burns" / f"{dex_stem(date_str)}-mints-burns.csv"
        df = pd.read_csv(path, parse_dates=["window_end"])
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end").set_index("window_end")
    out = out.rename(columns={
        "n_mints":        "dex_lp_n_mints",
        "n_burns":        "dex_lp_n_burns",
        "mint_liq":       "dex_lp_mint_liq_raw",
        "burn_liq":       "dex_lp_burn_liq_raw",
        "mint_amount0":   "dex_lp_mint_amount0_usdc",
        "mint_amount1":   "dex_lp_mint_amount1_usdt",
        "mint_usd":       "dex_lp_mint_usd",
        "burn_amount0":   "dex_lp_burn_amount0_usdc",
        "burn_amount1":   "dex_lp_burn_amount1_usdt",
        "burn_usd":       "dex_lp_burn_usd",
        "net_liq_change": "dex_lp_net_liq_change",
    })
    return out


# ==============================================================================
# CEX LOADERS
# ==============================================================================

def load_all_cex_klines():
    """
    Binance 5-min OHLCV klines. Timestamps are Unix microseconds.
    Base asset = USDC, Quote asset = USDT.
    cex_taker_buy_sell_ratio > 0.5 => net buying pressure on USDC.
    """
    frames = []
    for date_str in cex_kline_dates():
        path = CEX_DIR / "klines_cex" / f"USDCUSDT-5m-{date_str}.csv"
        df = pd.read_csv(path)
        # Binance window_end is the last microsecond of the candle (e.g. 00:04:59.999999).
        # Use window_start + 5min to get a clean 00:05:00 boundary instead.
        df["window_end"] = pd.to_datetime(
            df["window_start"], unit="us", utc=True
        ).dt.tz_localize(None) + FIVE_MIN
        for col in ["open", "high", "low", "close", "volume",
                    "quote_asset_volume",
                    "taker_buy_base_asset_volume",
                    "taker_buy_quote_asset_volume"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end").set_index("window_end")
    out = out.drop(columns=["window_start", "ignore"], errors="ignore")

    taker_buy  = out["taker_buy_base_asset_volume"]
    total_base = out["volume"]

    out = out.rename(columns={
        "open":                        "cex_price_open",
        "high":                        "cex_price_high",
        "low":                         "cex_price_low",
        "close":                       "cex_price_close",
        "volume":                      "cex_volume_usdc",
        "quote_asset_volume":          "cex_volume_usdt",
        "trades":                      "cex_n_trades",
        "taker_buy_base_asset_volume": "cex_taker_buy_usdc",
        "taker_buy_quote_asset_volume":"cex_taker_buy_usdt",
    })
    out["cex_taker_sell_usdc"]       = total_base - taker_buy
    out["cex_taker_buy_sell_ratio"]  = (taker_buy / total_base).where(total_base > 0)
    return out


def load_all_cex_orderbook():
    """
    Binance 5-min orderbook snapshots.
    spread = best_ask - best_bid (in USDT).
    imbalance = (bid_depth - ask_depth) / (bid_depth + ask_depth):
      +1 = only bids, -1 = only asks.
    bid_depth = total USDT posted on bid side within snapshot range.
    ask_depth = total USDC posted on ask side.
    """
    frames = []
    for date_str in orderbook_dates():
        path = CEX_DIR / "orderbook" / f"{dex_stem(date_str)}-orderbook.csv"
        df = pd.read_csv(path, parse_dates=["window_end"])
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True).sort_values("window_end").set_index("window_end")
    out = out.rename(columns={
        "n_seconds":      "cex_ob_n_seconds",
        "spread_mean":    "cex_ob_spread_mean",
        "spread_std":     "cex_ob_spread_std",
        "imbalance_mean": "cex_ob_imbalance_mean",
        "imbalance_std":  "cex_ob_imbalance_std",
        "bid_depth_mean": "cex_ob_bid_depth_mean",
        "bid_depth_max":  "cex_ob_bid_depth_max",
        "bid_depth_min":  "cex_ob_bid_depth_min",
        "ask_depth_mean": "cex_ob_ask_depth_mean",
        "ask_depth_max":  "cex_ob_ask_depth_max",
        "ask_depth_min":  "cex_ob_ask_depth_min",
    })
    return out


# ==============================================================================
# DUNE LOADERS  (activate with DUNE_ENABLED = True)
# ==============================================================================

def _dune_ts(series):
    """Parse Dune export timestamp: '2026-03-06 00:05:00.000 UTC' -> datetime."""
    return pd.to_datetime(series.str.replace(r"\.000 UTC$", "", regex=True))

def load_dune_gas_price():
    """
    Query 6763557 - Ethereum Gas Price Time Series
    Save export as: D:/data/dune_ONCHAIN/q3_gas_price.csv
    Key output columns:
      dune_gas_base_fee_gwei         avg base fee per block per minute (protocol-set)
      dune_gas_tip_p50_gwei          median priority fee (miner tip, user-set)
      dune_gas_tip_spread_gwei       p90-p10 priority fee spread (urgency dispersion)
      dune_gas_effective_gwei        approx total gas cost = base + median tip
      dune_gas_tx_count              number of transactions in the minute
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q3_gas_price.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping gas price")
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    df = df.set_index("window_end")
    rename = {
        "avg_base_fee_gwei":         "dune_gas_base_fee_gwei",
        "median_base_fee_gwei":      "dune_gas_base_fee_median_gwei",
        "min_base_fee_gwei":         "dune_gas_base_fee_min_gwei",
        "max_base_fee_gwei":         "dune_gas_base_fee_max_gwei",
        "priority_fee_p10_gwei":     "dune_gas_tip_p10_gwei",
        "priority_fee_p50_gwei":     "dune_gas_tip_p50_gwei",
        "priority_fee_p80_gwei":     "dune_gas_tip_p80_gwei",
        "gas_price_p10_gwei":        "dune_gas_price_p10_gwei",
        "gas_price_p50_gwei":        "dune_gas_price_p50_gwei",
        "gas_price_p80_gwei":        "dune_gas_price_p80_gwei",
        "approx_effective_gas_gwei": "dune_gas_effective_gwei",
        "tx_count":                  "dune_gas_tx_count",
        "block_count":               "dune_gas_block_count",
    }
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})


def load_dune_gas_used():
    """
    Query 6763559 - Ethereum Gas Used per Block & Block Utilization
    Save export as: D:/data/dune_ONCHAIN/q4_gas_used.csv
    Key output columns:
      dune_block_utilization         avg gas_used / gas_limit (>0.5 => base fee rises)
      dune_block_pct_above_target    fraction of blocks > 50% full (EIP-1559 pressure)
      dune_block_pct_near_full       fraction of blocks > 90% full (severe congestion)
      dune_block_blob_gas_mean       EIP-4844 blob gas (post-Dencun, March 2024+)
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q4_gas_used.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping gas used")
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    df = df.set_index("window_end")
    rename = {
        "block_count":             "dune_block_count",
        "avg_gas_used":            "dune_block_gas_used_mean",
        "min_gas_used":            "dune_block_gas_used_min",
        "max_gas_used":            "dune_block_gas_used_max",
        "median_gas_used":         "dune_block_gas_used_median",
        "avg_gas_limit":           "dune_block_gas_limit_mean",
        "avg_utilization":         "dune_block_utilization",
        "median_utilization":      "dune_block_utilization_median",
        "max_utilization":         "dune_block_utilization_max",
        "pct_blocks_above_target": "dune_block_pct_above_target",
        "pct_blocks_near_full":    "dune_block_pct_near_full",
        "avg_base_fee_gwei":       "dune_block_base_fee_gwei",
        "avg_blob_gas_used":       "dune_block_blob_gas_mean",
        "avg_block_size_bytes":    "dune_block_size_bytes_mean",
    }
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})


def load_dune_mempool():
    """
    Query 6763560 - Ethereum Mempool Congestion Proxies
    Save export as: D:/data/dune_ONCHAIN/q5_mempool_congestion.csv
    Key output columns:
      dune_mempool_fill_ratio        avg gas_used/gas_limit (congestion proxy)
      dune_mempool_base_fee_change   pct change in base fee minute-over-minute
      dune_mempool_tip_spread_gwei   p90-p10 tip spread (urgency dispersion)
      dune_mempool_congestion_score  composite 0-1 score:
                                       40% fill + 30% base fee rising + 30% tip spread
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q5_mempool_congestion.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping mempool")
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    df = df.set_index("window_end")
    rename = {
        "block_count":              "dune_mempool_block_count",
        "tx_count":                 "dune_mempool_tx_count",
        "avg_fill_ratio":           "dune_mempool_fill_ratio",
        "max_fill_ratio":           "dune_mempool_fill_ratio_max",
        "pct_blocks_near_full":     "dune_mempool_pct_near_full",
        "pct_blocks_above_target":  "dune_mempool_pct_above_target",
        "avg_base_fee_gwei":        "dune_mempool_base_fee_gwei",
        "base_fee_pct_change":      "dune_mempool_base_fee_change",
        "priority_fee_median_gwei": "dune_mempool_tip_median_gwei",
        "priority_fee_p80_gwei":    "dune_mempool_tip_p80_gwei",
        "priority_fee_spread_gwei": "dune_mempool_tip_spread_gwei",
        "avg_tx_gas_limit":         "dune_mempool_tx_gas_limit_mean",
        "avg_tx_gas_used":          "dune_mempool_tx_gas_used_mean",
        "congestion_score":         "dune_mempool_congestion_score",
    }
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})


def load_dune_supply_changes():
    """
    Query 6763561 - USDC/USDT True Mints & Burns (Protocol Supply Changes)
    Save export as: D:/data/dune_ONCHAIN/q6_supply_changes.csv
    Raw format: long table (one row per token x event_type per minute).
    This loader pivots to wide format (one row per minute).
    Key output columns (examples):
      dune_supply_usdc_mint_amount   USDC minted by Circle in this window
      dune_supply_usdc_burn_amount   USDC burned by Circle in this window
      dune_supply_usdt_mint_amount   USDT issued by Tether
      dune_supply_usdt_burn_amount   USDT redeemed by Tether
      dune_supply_usdc_net_delta     cumulative USDC supply change
      dune_supply_usdt_net_delta     cumulative USDT supply change
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q6_supply_changes.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping supply changes")
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    # Pivot long -> wide
    pivot = df.pivot_table(
        index="window_end",
        columns=["token", "event_type"],
        values=["total_token_amount", "event_count", "cumulative_net_supply_delta"],
        aggfunc="sum",
    )
    pivot.columns = [
        f"dune_supply_{tok.lower()}_{etype}_{metric}"
        for metric, tok, etype in pivot.columns
    ]
    return pivot


def load_dune_cex_flows():
    """
    Query 6763555 - USDC/USDT CEX Inflows & Outflows (Ethereum)
    Save export as: D:/data/dune_ONCHAIN/q2_cex_flows.csv
    Raw format: one row per (minute, cex_name, token, flow_type).
    This loader aggregates across all CEX names and pivots to wide.
    Key output columns (examples):
      dune_flows_usdc_inflow_usd     total USDC deposited to all CEXes in this window
      dune_flows_usdc_outflow_usd    total USDC withdrawn from all CEXes
      dune_flows_usdt_inflow_usd     total USDT deposited to all CEXes
      dune_flows_usdt_outflow_usd    total USDT withdrawn from all CEXes
    Note: inflow = transfer TO CEX hot wallet; outflow = transfer FROM CEX.
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q2_cex_flows.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping CEX flows")
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    # Lowercase flow_type so pivot column names are consistent (Inflow -> inflow)
    df["flow_type"] = df["flow_type"].str.lower()
    agg = df.groupby(["window_end", "token_symbol", "flow_type"], as_index=False).agg(
        total_usd      =("total_usd",       "sum"),
        transfer_count =("transfer_count",  "sum"),
        net_usd_flow   =("net_usd_flow",    "sum"),
    )
    pivot = agg.pivot_table(
        index="window_end",
        columns=["token_symbol", "flow_type"],
        values=["total_usd", "transfer_count"],
        aggfunc="sum",
    )
    pivot.columns = [
        f"dune_flows_{tok.lower()}_{ftype}_{metric}"
        for metric, tok, ftype in pivot.columns
    ]
    return pivot


def load_dune_whale_transfers():
    """
    Query 6763552 - Whale Transfers (>= $1M USD by default)
    Save export as: D:/data/dune_ONCHAIN/q1_whale_transfers.csv
    Raw format: one row per whale transfer event.
    This loader aggregates to (minute, token, flow_direction) and pivots.
    Key output columns (examples):
      dune_whale_usdc_cex_inflow_usd    large USDC transfers INTO CEXes
      dune_whale_usdc_cex_outflow_usd   large USDC transfers OUT of CEXes
      dune_whale_usdt_non_cex_usd       large USDT transfers between non-CEX wallets
    flow_direction values: cex_inflow, cex_outflow, cex_to_cex, non_cex
    """
    if not DUNE_ENABLED:
        return pd.DataFrame()
    path = DUNE_DIR / "q1_whale_transfers.csv"
    if not path.exists():
        print(f"  [DUNE] WARNING: {path} not found -- skipping whale transfers")
        return pd.DataFrame()
    # CSV is already aggregated per (window_end, token, flow_direction)
    df = pd.read_csv(path)
    df["window_end"] = _dune_ts(df["window_end"])
    pivot = df.pivot_table(
        index="window_end",
        columns=["token", "flow_direction"],
        values=["total_usd", "transfer_count"],
        aggfunc="sum",
    )
    pivot.columns = [
        f"dune_whale_{tok.lower()}_{direction}_{metric}"
        for metric, tok, direction in pivot.columns
    ]
    return pivot


# ==============================================================================
# MERGE
# ==============================================================================

def build_5min():
    sources = [
        ("DEX klines",         load_all_dex_klines),
        ("DEX swaps raw",      load_all_dex_swaps_raw),
        ("DEX pool state",     load_all_dex_pool),
        ("DEX ticks",          load_all_dex_ticks),
        ("DEX mints/burns",    load_all_dex_mints_burns),
        ("CEX klines",         load_all_cex_klines),
        ("CEX orderbook",      load_all_cex_orderbook),
        ("DUNE gas price",     load_dune_gas_price),
        ("DUNE gas used",      load_dune_gas_used),
        ("DUNE mempool",       load_dune_mempool),
        ("DUNE supply chg",    load_dune_supply_changes),
        ("DUNE CEX flows",     load_dune_cex_flows),
        ("DUNE whale xfers",   load_dune_whale_transfers),
    ]
    loaded = []
    for name, loader in sources:
        print(f"  Loading {name}...")
        df = loader()
        if not df.empty:
            print(f"    {len(df):>5} rows  x  {len(df.columns):>3} cols")
            loaded.append(df)
        else:
            print(f"    (empty / skipped)")

    if not loaded:
        raise RuntimeError("No data loaded at all -- check paths.")

    merged = loaded[0]
    for df in loaded[1:]:
        merged = merged.join(df, how="outer")
    return merged.sort_index()


# ==============================================================================
# EXCEL OUTPUT
# ==============================================================================

GROUPS = [
    # (prefix_list,                header_hex,  data_hex,  legend_label)
    (["cex_price_"],               "1F4E79", "D6E4F0", "CEX Price  --  Binance OHLC  (USDC/USDT)"),
    (["cex_volume_", "cex_n_",
      "cex_taker_"],               "2E75B6", "DAEEF3", "CEX Volume & Trade Flow  (base=USDC, quote=USDT)"),
    (["cex_ob_"],                  "4472C4", "DAE3F3", "CEX Orderbook  --  spread, depth, imbalance"),
    (["dex_klines_price_"],        "1E5631", "C6EFCE", "DEX Price  --  Uniswap v3 OHLC (from klines)"),
    (["dex_klines_volume_",
      "dex_klines_n_",
      "dex_klines_imbalance",
      "dex_klines_large_"],        "375623", "E2EFDA", "DEX Volume, Swap counts, Imbalance, Large trades"),
    (["dex_swaps_"],               "4CAF50", "EBF5EB", "DEX Raw Swaps  --  wallet counts, tick range, size"),
    (["dex_pool_"],                "00695C", "E0F2F1", "DEX Pool State  --  TVL, liquidity, price (sqrtPrice)"),
    (["dex_ticks_"],               "7F4700", "FFF2CC", "DEX Tick Distribution  --  active ticks, liq above/below"),
    (["dex_lp_"],                  "4B0082", "EAD1F7", "DEX LP Events  --  Mints (add liq) & Burns (remove liq)"),
    (["dune_gas_"],                "B8001F", "FFE3E3", "DUNE  --  Ethereum Gas Metrics  [activate: DUNE_ENABLED=True]"),
    (["dune_block_"],              "CC2200", "FFD6CC", "DUNE  --  Block Utilization & Gas Used  [skeleton]"),
    (["dune_mempool_"],            "DD3300", "FFDDD6", "DUNE  --  Mempool Congestion Score  [skeleton]"),
    (["dune_supply_"],             "8B4513", "FFDEAD", "DUNE  --  USDC/USDT On-chain Supply Changes  [skeleton]"),
    (["dune_flows_"],              "996633", "FFF0CC", "DUNE  --  On-chain CEX Inflow/Outflow  [skeleton]"),
    (["dune_whale_"],              "663399", "F0E6FF", "DUNE  --  Whale Transfers (>=1M USD)  [skeleton]"),
]
DATE_HDR_HEX = "404040"


def get_group(col):
    for prefixes, hdr, data, _ in GROUPS:
        if any(col.startswith(p) for p in prefixes):
            return hdr, data
    return "808080", "F2F2F2"


def write_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "5-min Data"

    cols = list(df.columns)

    # Header row
    def hdr_cell(col_i, value, fill_hex):
        c = ws.cell(1, col_i, value)
        c.fill      = PatternFill("solid", fgColor=fill_hex)
        c.font      = Font(bold=True, color="FFFFFF", size=8)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hdr_cell(1, "window_end (UTC)", DATE_HDR_HEX)
    for ci, col in enumerate(cols, start=2):
        fill_hex, _ = get_group(col)
        hdr_cell(ci, col, fill_hex)

    ws.row_dimensions[1].height = 90

    # Data -- use dataframe_to_rows for speed (thousands of rows)
    df_out = df.reset_index()
    df_out["window_end"] = df_out["window_end"].astype(str)
    for ri, row_data in enumerate(dataframe_to_rows(df_out, index=False, header=False), start=2):
        for ci, val in enumerate(row_data, start=1):
            if isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(ri, ci, val)

    # Column widths
    ws.column_dimensions["A"].width = 20   # window_end
    for ci, col in enumerate(cols, start=2):
        w = max(len(w) for w in col.split("_")) + 3
        ws.column_dimensions[get_column_letter(ci)].width = max(w, 10)

    ws.freeze_panes = "B2"

    # Legend sheet
    wl = wb.create_sheet("Legend & Guide")
    wl.column_dimensions["A"].width = 6
    wl.column_dimensions["B"].width = 55
    wl.column_dimensions["C"].width = 75

    def lhdr(row, *vals):
        for ci, v in enumerate(vals, start=1):
            c = wl.cell(row, ci, v)
            c.fill = PatternFill("solid", fgColor=DATE_HDR_HEX)
            c.font = Font(bold=True, color="FFFFFF", size=10)

    ri = 1
    lhdr(ri, "Color", "Column Group", "Column Prefixes"); ri += 1

    for prefixes, hdr_hex, data_hex, label in GROUPS:
        sw = wl.cell(ri, 1, "  ")
        sw.fill = PatternFill("solid", fgColor=hdr_hex)
        lb = wl.cell(ri, 2, label)
        lb.fill = PatternFill("solid", fgColor=data_hex)
        lb.font = Font(bold=True, size=10)
        wl.cell(ri, 3, "  |  ".join(prefixes))
        ri += 1

    ri += 1
    lhdr(ri, "", "Suffix Guide", "Meaning"); ri += 1
    for sfx, meaning in [
        ("_open",         "First value of the 5-min window (price at window start)"),
        ("_high / _low",  "Max / min value within the 5-min window"),
        ("_close",        "Last value -- i.e. the price AT window_end timestamp"),
        ("_mean",         "Average across samples within the 5-min window"),
        ("_std",          "Standard deviation within the 5-min window"),
        ("_max / _min",   "Extreme values observed within the 5-min window"),
        ("_n_total",      "Count of events within the 5-min window"),
        ("_usd",          "USD-equivalent value (using USDT as numeraire)"),
        ("_usdc",         "Amount denominated in USDC"),
        ("_usdt",         "Amount denominated in USDT"),
        ("_liq_raw",      "Uniswap v3 raw liquidity units (L = sqrt(xy), NOT USD)"),
        ("_gwei",         "Gas price in Gwei (1 Gwei = 1e-9 ETH)"),
    ]:
        wl.cell(ri, 2, sfx).font = Font(bold=True, size=10)
        wl.cell(ri, 3, meaning).alignment = Alignment(wrap_text=True)
        ri += 1

    ri += 1
    lhdr(ri, "", "Key Concepts", "Explanation"); ri += 1
    for key, explanation in [
        ("window_end",
         "The timestamp at which the 5-min window CLOSES. "
         "e.g. 00:05:00 UTC = data from 00:00:00 to 00:05:00. "
         "Price columns (close) = exact price AT this moment."),
        ("dex_klines_imbalance",
         "net_amount0 / |amount0| per 5-min window. "
         "+1 = all USDC flowed INTO pool (traders swapping USDC for USDT). "
         "-1 = all USDC flowed OUT (traders swapping USDT for USDC). "
         "Proxy for directional flow pressure on the DEX."),
        ("cex_taker_buy_sell_ratio",
         "taker_buy_volume / total_volume. >0.5 = net buying pressure on USDC "
         "(market orders buying USDC, paying USDT). <0.5 = net selling."),
        ("cex_ob_imbalance",
         "(bid_depth - ask_depth) / (bid_depth + ask_depth). "
         "+1 = only bids (strong buying interest). -1 = only asks (selling interest)."),
        ("dex_pool_price vs dex_klines_price",
         "Both measure USDC/USDT ratio but from different sources: "
         "pool price = from sqrtPriceX96 (pool contract state, end-of-window snapshot). "
         "klines price = from individual swap events (OHLC within window). "
         "Should be very close but not identical."),
        ("dex_ticks_net_liq_above/below",
         "Net liquidity in ticks ABOVE (above current price) vs BELOW. "
         "Imbalance indicates where LPs are positioned. "
         "More liq above = LPs expect upward price moves."),
        ("dune_mempool_congestion_score",
         "Composite 0-1 score: "
         "40% x avg_fill_ratio + 30% x (1 if base fee rising) + 30% x min(tip_spread/20, 1). "
         "Higher = more congested mempool. Useful as a gas/latency risk proxy."),
        ("DUNE columns (all NaN)",
         "Dune data is not yet loaded. Set DUNE_ENABLED=True and place CSV exports "
         "in D:/data/dune_ONCHAIN/ with the filenames specified in the loader docstrings."),
    ]:
        wl.cell(ri, 2, key).font = Font(bold=True, size=10)
        wl.cell(ri, 3, explanation).alignment = Alignment(wrap_text=True)
        wl.row_dimensions[ri].height = 36
        ri += 1

    wb.save(OUT_XLSX)
    df.to_csv(OUT_CSV)
    print(f"\nSaved Excel : {OUT_XLSX}")
    print(f"Saved CSV   : {OUT_CSV}")


# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    print("Building 5-min aggregated dataset...")
    df = build_5min()

    print(f"\n  Total 5-min rows : {len(df)}")
    print(f"  Time range       : {df.index[0]}  ->  {df.index[-1]}")
    print(f"  Total columns    : {len(df.columns)}")

    print(f"\n  Column groups:")
    for prefixes, _, _, label in GROUPS:
        matching = [c for c in df.columns if any(c.startswith(p) for p in prefixes)]
        if matching:
            print(f"    {label:<65} {len(matching):>2} cols")

    print(f"\n  Data coverage (non-NaN rows per group out of {len(df)} total):")
    for prefixes, _, _, label in GROUPS:
        group_cols = [c for c in df.columns if any(c.startswith(p) for p in prefixes)]
        if group_cols:
            n = df[group_cols].notna().any(axis=1).sum()
            print(f"    {label:<65} {n:>5} rows  ({n/len(df)*100:5.1f}%)")

    write_excel(df)
    print("\nDone.")
